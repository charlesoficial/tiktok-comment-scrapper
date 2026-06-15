import csv
import json

import tiktok_exporter.exporter as exporter
from tiktok_exporter.models import Comment


class FakeClient:
    def iter_comments(self, video_id, *, limit, include_replies):
        reply = Comment(
            id="reply-1",
            username="reply_user",
            nickname="Reply User",
            text="reply text",
            created_at="2026-06-15T10:01:00",
            avatar_url="",
            like_count=0,
            reply_count=0,
        )
        comment = Comment(
            id="comment-1",
            username="comment_user",
            nickname="Comment User",
            text="comment text",
            created_at="2026-06-15T10:00:00",
            avatar_url="",
            like_count=3,
            reply_count=1,
            replies=[reply] if include_replies else [],
        )
        return [comment][:limit]

    def get_video_meta(self, video_id):
        return "caption", f"https://www.tiktok.com/@user/video/{video_id}"


def test_export_comments_writes_json(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        include_replies=True,
    )

    assert len(files) == 1
    data = json.loads(files[0].read_text(encoding="utf-8"))
    assert data["video_id"] == "7418294751977327878"
    assert data["comment_count"] == 1
    assert data["comments"][0]["replies"][0]["id"] == "reply-1"


def test_export_comments_writes_csv(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        include_replies=True,
        file_format="csv",
    )

    assert files[0].suffix == ".csv"
    with files[0].open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.DictReader(file))

    assert len(rows) == 2
    assert rows[0]["id"] == "comment-1"
    assert rows[0]["is_reply"] == "False"
    assert rows[1]["id"] == "reply-1"
    assert rows[1]["parent_id"] == "comment-1"
    assert rows[1]["is_reply"] == "True"


def test_export_comments_writes_both_formats(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        file_format="both",
    )

    assert {file.suffix for file in files} == {".json", ".csv"}


def test_client_reuses_first_page_metadata_without_extra_request():
    from tiktok_exporter.client import TikTokClient

    client = TikTokClient()
    calls: list[dict] = []

    def fake_get(endpoint, params):
        calls.append(params)
        if params["cursor"] == 0:
            return {
                "comments": [
                    {
                        "cid": "c1",
                        "text": "hi",
                        "create_time": 1710000000,
                        "digg_count": 0,
                        "reply_comment_total": 0,
                        "user": {"unique_id": "u", "nickname": "U"},
                        "share_info": {"title": "caption", "url": "https://t/v/1"},
                    }
                ],
                "has_more": False,
                "cursor": 50,
            }
        return {"comments": [], "has_more": False}

    client._get = fake_get  # type: ignore[method-assign]

    list(client.iter_comments("1", limit=10, include_replies=False))
    caption, video_url = client.get_video_meta("1")

    assert caption == "caption"
    assert video_url == "https://t/v/1"
    # apenas a pagina de comentarios deve ter sido buscada, sem chamada extra de meta
    assert len(calls) == 1


def test_export_comments_writes_txt(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        include_replies=True,
        file_format="txt",
    )

    assert files[0].suffix == ".txt"
    content = files[0].read_text(encoding="utf-8")
    assert "comment text" in content
    assert "@comment_user" in content
    assert "reply text" in content


def test_export_comments_all_formats(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        file_format="all",
    )

    assert {file.suffix for file in files} == {".json", ".csv", ".txt", ".xlsx"}


def test_export_comments_writes_xlsx(monkeypatch, tmp_path):
    monkeypatch.setattr(exporter, "TikTokClient", FakeClient)

    files = exporter.export_comments(
        "7418294751977327878",
        limit=1,
        output_dir=tmp_path,
        include_replies=True,
        file_format="xlsx",
    )

    assert files[0].suffix == ".xlsx"

    from openpyxl import load_workbook

    workbook = load_workbook(files[0])
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == tuple(exporter.TABLE_COLUMNS)
    assert rows[1][2] == "comment-1"
    assert rows[2][2] == "reply-1"
