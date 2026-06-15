from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Callable

from .client import TikTokClient
from .models import Comment
from .models import ExportResult
from .utils import extract_video_id

ProgressCallback = Callable[[str], None]

TABLE_COLUMNS = [
    "video_id",
    "parent_id",
    "id",
    "username",
    "nickname",
    "text",
    "created_at",
    "avatar_url",
    "like_count",
    "reply_count",
    "is_reply",
]


def _report(progress: ProgressCallback | None, message: str) -> None:
    if progress is not None:
        progress(message)


def export_comments(
    video: str,
    *,
    limit: int | None,
    output_dir: str | Path,
    include_replies: bool = True,
    pretty: bool = True,
    file_format: str = "json",
    progress: ProgressCallback | None = None,
) -> list[Path]:
    video_id = extract_video_id(video)
    _report(progress, f"Coletando comentarios do video {video_id}...")

    client = TikTokClient()
    comments: list[Comment] = []
    for comment in client.iter_comments(
        video_id,
        limit=limit,
        include_replies=include_replies,
    ):
        comments.append(comment)
        if len(comments) % 10 == 0:
            _report(progress, f"{len(comments)} comentarios coletados...")

    _report(progress, f"Total de {len(comments)} comentarios coletados.")

    caption, video_url = client.get_video_meta(video_id)

    result = ExportResult(
        video_id=video_id,
        caption=caption,
        video_url=video_url,
        comment_count=len(comments),
        comments=comments,
    )

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    files: list[Path] = []

    if file_format in {"json", "both", "all"}:
        files.append(_write_json(result, destination, pretty=pretty))

    if file_format in {"csv", "both", "all"}:
        files.append(_write_csv(result, destination))

    if file_format in {"txt", "all"}:
        files.append(_write_txt(result, destination))

    if file_format in {"xlsx", "all"}:
        files.append(_write_xlsx(result, destination))

    return files


def _write_json(result: ExportResult, output_dir: Path, *, pretty: bool) -> Path:
    file_path = output_dir / f"{result.video_id}.json"
    file_path.write_text(
        json.dumps(result.to_dict(), ensure_ascii=False, indent=2 if pretty else None),
        encoding="utf-8",
    )
    return file_path


def _write_csv(result: ExportResult, output_dir: Path) -> Path:
    file_path = output_dir / f"{result.video_id}.csv"
    with file_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=TABLE_COLUMNS)
        writer.writeheader()
        for row in _iter_csv_rows(result.video_id, result.comments):
            writer.writerow(row)
    return file_path


def _iter_csv_rows(video_id: str, comments: list[Comment]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for comment in comments:
        rows.append(_comment_to_row(video_id, comment, parent_id="", is_reply=False))
        for reply in comment.replies:
            rows.append(_comment_to_row(video_id, reply, parent_id=comment.id, is_reply=True))
    return rows


def _comment_to_row(
    video_id: str,
    comment: Comment,
    *,
    parent_id: str,
    is_reply: bool,
) -> dict[str, object]:
    return {
        "video_id": video_id,
        "parent_id": parent_id,
        "id": comment.id,
        "username": comment.username,
        "nickname": comment.nickname,
        "text": comment.text,
        "created_at": comment.created_at,
        "avatar_url": comment.avatar_url,
        "like_count": comment.like_count,
        "reply_count": comment.reply_count,
        "is_reply": is_reply,
    }


def _write_txt(result: ExportResult, output_dir: Path) -> Path:
    file_path = output_dir / f"{result.video_id}.txt"
    lines: list[str] = []

    lines.append(f"Video: {result.video_id}")
    if result.caption:
        lines.append(f"Legenda: {result.caption}")
    if result.video_url:
        lines.append(f"URL: {result.video_url}")
    lines.append(f"Total de comentarios: {result.comment_count}")
    lines.append("=" * 60)
    lines.append("")

    for index, comment in enumerate(result.comments, start=1):
        lines.extend(_comment_to_lines(index, comment))
        lines.append("")

    file_path.write_text("\n".join(lines), encoding="utf-8")
    return file_path


def _comment_to_lines(index: int, comment: Comment) -> list[str]:
    header = f"{index}. {comment.nickname} (@{comment.username})"
    meta = f"   {comment.created_at} | {comment.like_count} likes"
    lines = [header, f"   {comment.text}", meta]

    for reply in comment.replies:
        lines.append(f"   |- {reply.nickname} (@{reply.username}): {reply.text}")
        lines.append(f"      {reply.created_at} | {reply.like_count} likes")

    return lines


def _write_xlsx(result: ExportResult, output_dir: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depende de instalacao externa
        raise RuntimeError(
            "exportar para xlsx requer o pacote openpyxl. Instale com: pip install openpyxl"
        ) from exc

    file_path = output_dir / f"{result.video_id}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "comentarios"

    sheet.append(TABLE_COLUMNS)
    for row in _iter_csv_rows(result.video_id, result.comments):
        sheet.append([row[column] for column in TABLE_COLUMNS])

    for index, column in enumerate(TABLE_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, len(column) + 2)
    sheet.freeze_panes = "A2"

    workbook.save(file_path)
    return file_path
